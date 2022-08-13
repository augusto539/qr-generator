from openpyxl import load_workbook

def Load_excel_info(xlsx_route: str):
    # cargo el libro de calculo
    book = load_workbook(xlsx_route)

    # declaro la forma que quiero que tengan los diccionaries para guardar la info
    data = {
        "ID": [],
        'Nombres': [],
        'Apellidos': []
    }
    # declaro la lista de diccionarios
    list_of_data = []

    # recorro las hojas del libro
    for sheet_name in book.sheetnames:
        # uso el nombre de la hoja del libro para inicializar el la variable "sheet" que es igual al contenido de la hoja de calculo
        sheet = book[sheet_name]
        # recorro las lineas de la hoja de calculo
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue
            else:
                # gurado la data en su respectiva parte del diccionario
                data['ID'] = row[0]
                data['Apellidos'] = row[2]
                data['Nombres'] = row[3]
            # agrego el nuevo diccionario a la lista
            # print(data)
            list_of_data.append(data)
            data = {
                "ID": [],
                'Nombres': [],
                'Apellidos': []
            }
    # regreso la lista con los datos    
    return list_of_data
